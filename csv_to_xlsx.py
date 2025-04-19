import pandas as pd
import openpyxl
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.axis import ChartLines
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

def main():
    # Path configuration
    csv_path = 'temp_csv_data.csv'
    excel_path = 'final_xlsx_data.xlsx'
    
    # Read CSV data
    df = pd.read_csv(csv_path)
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Create data sheet and add data
    data_sheet = wb.active
    data_sheet.title = 'Feedback Data'
    
    # Add data to the sheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            data_sheet.cell(row=r_idx, column=c_idx, value=value)
    
    # Define table range
    data_rows = len(df) + 1  # Add 1 for header row
    data_cols = len(df.columns)
    table_range = f"A1:{get_column_letter(data_cols)}{data_rows}"
    
    # Create and add the table
    table = Table(displayName="FeedbackTable", ref=table_range)
    style = TableStyleInfo(
        name="TableStyleMedium9", 
        showFirstColumn=False,
        showLastColumn=False, 
        showRowStripes=True, 
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    data_sheet.add_table(table)
    
    # Create charts sheet
    chart_sheet = wb.create_sheet('Charts')
    
    # Get unique values for each category
    departments = ['highschool', 'college']
    ratings = ['Needs Improvement', 'Average', 'Good']
    users = ['student', 'faculty', 'visitor', 'ntp']
    
    # Create Department Summary using COUNTIF formulas
    chart_sheet['A1'] = 'Department Summary'
    chart_sheet['A2'] = 'Department'
    chart_sheet['B2'] = 'Count'
    
    for i, dept in enumerate(departments):
        row = i + 3
        chart_sheet[f'A{row}'] = dept
        chart_sheet[f'B{row}'] = f'=COUNTIF(\'Feedback Data\'!B:B,"{dept}")'
    
    dept_range = f"A2:B{2+len(departments)}"
    dept_table = Table(displayName="DeptSummary", ref=dept_range)
    dept_table.tableStyleInfo = style
    chart_sheet.add_table(dept_table)
    
    # Create Rating Summary using COUNTIF formulas
    chart_sheet['D1'] = 'Rating Summary'
    chart_sheet['D2'] = 'Rating'
    chart_sheet['E2'] = 'Count'
    
    for i, rating in enumerate(ratings):
        row = i + 3
        chart_sheet[f'D{row}'] = rating
        chart_sheet[f'E{row}'] = f'=COUNTIF(\'Feedback Data\'!C:C,"{rating}")'
    
    rating_range = f"D2:E{2+len(ratings)}"
    rating_table = Table(displayName="RatingSummary", ref=rating_range)
    rating_table.tableStyleInfo = style
    chart_sheet.add_table(rating_table)
    
    # Create User Summary using COUNTIF formulas
    chart_sheet['G1'] = 'User Summary'
    chart_sheet['G2'] = 'User'
    chart_sheet['H2'] = 'Count'
    
    for i, user in enumerate(users):
        row = i + 3
        chart_sheet[f'G{row}'] = user
        chart_sheet[f'H{row}'] = f'=COUNTIF(\'Feedback Data\'!A:A,"{user}")'
    
    user_range = f"G2:H{2+len(users)}"
    user_table = Table(displayName="UserSummary", ref=user_range)
    user_table.tableStyleInfo = style
    chart_sheet.add_table(user_table)
    
    # Create Department Pie Chart
    pie_dept = PieChart()
    pie_dept.title = "Feedback by Department"
    pie_dept.legend.position = 'b'  # Position legend at bottom
    
    # Define data range for department pie chart
    data = Reference(chart_sheet, min_col=2, min_row=3, max_row=2+len(departments))
    categories = Reference(chart_sheet, min_col=1, min_row=3, max_row=2+len(departments))
    
    pie_dept.add_data(data, titles_from_data=False)
    pie_dept.set_categories(categories)
    
    # Configure data labels
    pie_dept.dataLabels = DataLabelList()
    pie_dept.dataLabels.showPercent = True
    pie_dept.dataLabels.showVal = False
    pie_dept.dataLabels.showCatName = True
    pie_dept.dataLabels.showSerName = False  # Don't show series name
    
    # Add chart to sheet - first chart starting at column A
    chart_sheet.add_chart(pie_dept, f"A10")
    
    # Create Rating Pie Chart
    pie_rating = PieChart()
    pie_rating.title = "Feedback by Rating"
    pie_rating.legend.position = 'b'  # Position legend at bottom
    
    # Define data range for rating pie chart
    data = Reference(chart_sheet, min_col=5, min_row=3, max_row=2+len(ratings))
    categories = Reference(chart_sheet, min_col=4, min_row=3, max_row=2+len(ratings))
    
    pie_rating.add_data(data, titles_from_data=False)
    pie_rating.set_categories(categories)
    
    # Configure data labels
    pie_rating.dataLabels = DataLabelList()
    pie_rating.dataLabels.showPercent = True
    pie_rating.dataLabels.showVal = False
    pie_rating.dataLabels.showCatName = True
    pie_rating.dataLabels.showSerName = False  # Don't show series name
    
    # Add chart to sheet - second chart starting at column J (which is position J)
    chart_sheet.add_chart(pie_rating, f"A26")
    
    # Create User Bar Chart
    bar_user = BarChart()
    bar_user.title = "Feedback by User Type"
    bar_user.y_axis.title = "Count"
    bar_user.x_axis.title = "User Type"
    bar_user.legend = None  # Remove legend entirely
    
    # Show axis lines
    bar_user.y_axis.delete = False  # Make sure y-axis is visible
    bar_user.x_axis.delete = False  # Make sure x-axis is visible
    
    # Define data range for user bar chart
    data = Reference(chart_sheet, min_col=8, min_row=3, max_row=2+len(users))
    categories = Reference(chart_sheet, min_col=7, min_row=3, max_row=2+len(users))
    
    bar_user.add_data(data, titles_from_data=False)
    bar_user.set_categories(categories)
    
    # Add chart to sheet - third chart starting at column S
    chart_sheet.add_chart(bar_user, f"J10")
    
    # Set chart sheet as the active sheet when opening the file
    wb.active = wb.index(chart_sheet)
    
    # Save the workbook
    wb.save(excel_path)
    
    print(f"Excel file with tables and charts created at {excel_path}")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"Error creating Excel file: {e}")
        # Ensure the script exits with an error code if there's an exception
        exit(1)